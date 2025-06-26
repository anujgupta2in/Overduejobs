import pandas as pd
import os
import re
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go

def process_csv_file(file):
    """Process a single CSV file and extract relevant information."""
    # Initialize variables with default values
    filename = "Unknown"
    formatted_date = "Unknown"
    
    try:
        # Extract date from filename using regex
        filename = file.name
        date_match = re.search(r'\b(\d{2})(\d{2})(\d{4})\b', filename)
        formatted_date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}" if date_match else "Unknown"
        
        # Read CSV file
        df = pd.read_csv(file)
        
        # Identify the Vessel column
        vessel_column = next((col for col in df.columns if 'vessel' in col.lower()), None)
        vessel_name = df[vessel_column].iloc[0] if vessel_column else "Vessel column not found"
        
        # Identify the Job Status column
        status_column = next((col for col in df.columns if 'status' in col.lower()), None)
        
        # Count total jobs
        job_count = len(df)
        
        # Count jobs with Job Status == 'New'
        if status_column:
            df[status_column] = df[status_column].astype(str).str.strip()
            new_job_count = (df[status_column] == 'New').sum()
        else:
            new_job_count = 0
            
        return {
            'File Name': filename,
            'Vessel Name': vessel_name,
            'Total Count of Jobs': job_count,
            'New Job Count': new_job_count,
            'Date Extracted from File Name': formatted_date
        }
    except Exception as e:
        return {
            'File Name': filename,
            'Vessel Name': 'Error',
            'Total Count of Jobs': 'Error',
            'New Job Count': 'Error',
            'Date Extracted from File Name': formatted_date
        }

def get_effective_date(file_name, today):
    try:
        # Extract date part from file name assuming format like "Ragnar 02032025"
        parts = file_name.split()
        for part in parts:
            if part.isdigit() and len(part) == 8:
                date_obj = datetime.strptime(part, "%d%m%Y")
                return date_obj
    except Exception as e:
        print(f"Date parsing error for file {file_name}: {e}")
    return today  # Default to today if parsing fails

def analyze_overdue_jobs(df):
    """Analyze overdue jobs and critical overdue jobs from a DataFrame.

    Returns a dictionary with overdue job metrics per individual file/record.
    """
    try:
        df_copy = df.copy()
        df_copy.columns = df_copy.columns.str.strip()

        file_results = []

        if 'Calculated Due Date' in df_copy.columns and 'Job Status' in df_copy.columns:
            df_copy['Calculated Due Date'] = pd.to_datetime(df_copy['Calculated Due Date'], format='%d-%m-%Y', errors='coerce')
            today = pd.to_datetime(datetime.today().date())

            if '_source_file' in df_copy.columns:
                files = df_copy['_source_file'].unique()
                file_col = '_source_file'
            elif 'File Name' in df_copy.columns:
                files = df_copy['File Name'].unique()
                file_col = 'File Name'
            else:
                files = ['Entire Dataset']
                df_copy['_file_id'] = 'Entire Dataset'
                file_col = '_file_id'

            for file_name in files:
                file_data = df_copy[df_copy[file_col] == file_name]

                # Use effective date based on filename, but fallback to today if file date == today
                # Fix: Strip extension from file name for accurate date extraction
                base_name = os.path.splitext(os.path.basename(str(file_name)))[0]
                file_date = get_effective_date(base_name, today)
                
                today_date = pd.to_datetime(datetime.today().date())
                effective_date = today_date if file_date.date() == today_date.date() else file_date

                overdue_jobs = file_data[
                    (file_data['Calculated Due Date'] <= effective_date) &
                    (file_data['Job Status'].astype(str).str.strip().str.lower().isin(['pending', 'in progress on board']))
                ]
                overdue_jobs_count = len(overdue_jobs)

                try:
                    if 'Unnamed: 0' in file_data.columns:
                        critical_overdue_jobs = file_data[
                            (file_data['Unnamed: 0'].astype(str).str.strip().str.lower() == 'c') &
                            (file_data['Calculated Due Date'] <= effective_date) &
                            (file_data['Job Status'].astype(str).str.strip().str.lower().isin(['pending', 'in progress on board']))
                        ]
                    else:
                        critical_col = next((col for col in file_data.columns if 'critical' in col.lower() or 'priority' in col.lower()), None)
                        if critical_col:
                            critical_overdue_jobs = file_data[
                                (file_data[critical_col].astype(str).str.strip().str.lower().isin(['c', 'critical', 'high', 'yes', 'true'])) &
                                (file_data['Calculated Due Date'] <= effective_date) &
                                (file_data['Job Status'].astype(str).str.strip().str.lower().isin(['pending', 'in progress on board']))
                            ]
                        else:
                            critical_overdue_jobs = pd.DataFrame()
                except Exception as e:
                    print(f"Error processing critical jobs for {file_name}: {str(e)}")
                    critical_overdue_jobs = pd.DataFrame()

                critical_overdue_jobs_count = len(critical_overdue_jobs)
                total_jobs = len(file_data)

                overdue_jobs_percentage = round((overdue_jobs_count / total_jobs) * 100, 2) if total_jobs else 0
                critical_overdue_jobs_percentage = round((critical_overdue_jobs_count / total_jobs) * 100, 2) if total_jobs else 0

                file_results.append({
                    'file_name': file_name,
                    'total_jobs': total_jobs,
                    'overdue_jobs_count': overdue_jobs_count,
                    'overdue_jobs_percentage': overdue_jobs_percentage,
                    'critical_overdue_jobs_count': critical_overdue_jobs_count,
                    'critical_overdue_jobs_percentage': critical_overdue_jobs_percentage,
                    'overdue_jobs': overdue_jobs,
                    'critical_overdue_jobs': critical_overdue_jobs
                })

            results_df = pd.DataFrame(file_results)
            total_all_jobs = results_df['total_jobs'].sum()
            total_overdue = results_df['overdue_jobs_count'].sum()
            total_critical = results_df['critical_overdue_jobs_count'].sum()

            overall_overdue_pct = round((total_overdue / total_all_jobs) * 100, 2) if total_all_jobs else 0
            overall_critical_pct = round((total_critical / total_all_jobs) * 100, 2) if total_all_jobs else 0

            all_overdue = pd.concat([result['overdue_jobs'] for result in file_results]) if file_results else pd.DataFrame()
            all_critical = pd.concat([result['critical_overdue_jobs'] for result in file_results]) if file_results else pd.DataFrame()

            return {
                'file_results': file_results,
                'overdue_jobs_count': total_overdue,
                'overdue_jobs_percentage': overall_overdue_pct,
                'critical_overdue_jobs_count': total_critical,
                'critical_overdue_jobs_percentage': overall_critical_pct,
                'total_jobs': total_all_jobs,
                'overdue_jobs': all_overdue,
                'critical_overdue_jobs': all_critical
            }

        else:
            return {
                'file_results': [],
                'overdue_jobs_count': 0,
                'overdue_jobs_percentage': 0,
                'critical_overdue_jobs_count': 0,
                'critical_overdue_jobs_percentage': 0,
                'total_jobs': 0,
                'overdue_jobs': pd.DataFrame(),
                'critical_overdue_jobs': pd.DataFrame()
            }

    except Exception as e:
        print(f"Error analyzing overdue jobs: {str(e)}")
        return {
            'file_results': [],
            'overdue_jobs_count': 0,
            'overdue_jobs_percentage': 0,
            'critical_overdue_jobs_count': 0,
            'critical_overdue_jobs_percentage': 0,
            'total_jobs': 0,
            'overdue_jobs': pd.DataFrame(),
            'critical_overdue_jobs': pd.DataFrame()
        }

def create_vessel_job_distribution_chart(df, overdue_data=None):
    """Create a bar chart showing job distribution across vessels for individual files.
    
    Args:
        df: DataFrame with vessel job data
        overdue_data: Optional dictionary with overdue jobs data
    """
    # Sort data by date to maintain chronological order
    df = df.sort_values('Date Extracted from File Name')
    
    fig = go.Figure()
    
    # Add total jobs bars
    fig.add_trace(go.Bar(
        name='Total Jobs',
        x=[f"{row['Vessel Name']} - {row['File Name']}" for _, row in df.iterrows()],
        y=df['Total Count of Jobs'],
        marker_color='#1E88E5'  # Blue for Total Jobs
    ))
    
    # Add new jobs bars
    fig.add_trace(go.Bar(
        name='New Jobs',
        x=[f"{row['Vessel Name']} - {row['File Name']}" for _, row in df.iterrows()],
        y=df['New Job Count'],
        marker_color='#4CAF50'  # Green for New Jobs
    ))
    
    # Add overdue jobs bars if data is provided
    if overdue_data and 'file_results' in overdue_data and overdue_data['file_results']:
        # Create mappings of file names to overdue and critical overdue counts
        overdue_map = {}
        critical_map = {}
        
        # Process each file result from the analysis
        for file_result in overdue_data['file_results']:
            file_name = file_result['file_name']
            overdue_map[file_name] = file_result['overdue_jobs_count']
            critical_map[file_name] = file_result['critical_overdue_jobs_count']
        
        # Create lists for overdue jobs per vessel-file combination
        overdue_jobs = []
        critical_overdue_jobs = []
        
        # Match overdue data with the vessel-file combinations
        for _, row in df.iterrows():
            file_name = row['File Name']
            # Try to find an exact match
            if file_name in overdue_map:
                overdue_jobs.append(overdue_map[file_name])
                critical_overdue_jobs.append(critical_map[file_name])
            else:
                # Try basename matching or partial matching
                matched = False
                file_basename = os.path.basename(file_name)
                
                # Try matching with just the filename
                if file_basename in overdue_map:
                    overdue_jobs.append(overdue_map[file_basename])
                    critical_overdue_jobs.append(critical_map[file_basename])
                    matched = True
                else:
                    # Try partial matching
                    for analysis_file in overdue_map:
                        if file_name in analysis_file or analysis_file in file_name:
                            overdue_jobs.append(overdue_map[analysis_file])
                            critical_overdue_jobs.append(critical_map[analysis_file])
                            matched = True
                            break
                
                # If no match found, add zeros
                if not matched:
                    overdue_jobs.append(0)
                    critical_overdue_jobs.append(0)
        
        # Add overdue jobs bars for each vessel-file
        if any(overdue_jobs):
            fig.add_trace(go.Bar(
                name='Overdue Jobs',
                x=[f"{row['Vessel Name']} - {row['File Name']}" for _, row in df.iterrows()],
                y=overdue_jobs,
                marker_color='#FF9800'  # Orange for Overdue Jobs
            ))
        
        # Add critical overdue jobs bars for each vessel-file
        if any(critical_overdue_jobs):
            fig.add_trace(go.Bar(
                name='Critical Overdue Jobs',
                x=[f"{row['Vessel Name']} - {row['File Name']}" for _, row in df.iterrows()],
                y=critical_overdue_jobs,
                marker_color='#F44336'  # Red for Critical Overdue Jobs
            ))
    
    # Update layout with improved readability
    fig.update_layout(
        title='Job Distribution by Vessel and File',
        xaxis_title='Vessel - File',
        yaxis_title='Number of Jobs',
        barmode='group',
        height=500,  # Increased height for better visibility
        showlegend=True,
        xaxis=dict(
            tickangle=45,  # Angled labels for better readability
            tickmode='array',
            ticktext=[f"{row['Vessel Name']}<br>{row['File Name']}" for _, row in df.iterrows()],
            tickvals=list(range(len(df)))
        ),
        margin=dict(b=150)  # Increased bottom margin for rotated labels
    )
    
    return fig

def create_jobs_timeline_chart(df, overdue_data=None):
    """Create a line chart showing job trends over time.
    
    Args:
        df: DataFrame with job data
        overdue_data: Optional dictionary with overdue jobs data
    """
    timeline_data = df.groupby('Date Extracted from File Name').agg({
        'Total Count of Jobs': 'sum',
        'New Job Count': 'sum'
    }).reset_index()
    
    # Sort by date
    timeline_data['Date Extracted from File Name'] = pd.to_datetime(timeline_data['Date Extracted from File Name'])
    timeline_data = timeline_data.sort_values('Date Extracted from File Name')
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=timeline_data['Date Extracted from File Name'],
        y=timeline_data['Total Count of Jobs'],
        name='Total Jobs',
        line=dict(color='#1E88E5', width=2)  # Blue for Total Jobs
    ))
    fig.add_trace(go.Scatter(
        x=timeline_data['Date Extracted from File Name'],
        y=timeline_data['New Job Count'],
        name='New Jobs',
        line=dict(color='#4CAF50', width=2)  # Green for New Jobs
    ))
    
    # Add overdue jobs to the timeline if data exists
    if overdue_data and 'file_results' in overdue_data and overdue_data['file_results']:
        # Group data by date
        date_to_file = {}
        for _, row in df.iterrows():
            date = pd.to_datetime(row['Date Extracted from File Name'])
            file_name = row['File Name']
            if date not in date_to_file:
                date_to_file[date] = []
            date_to_file[date].append(file_name)
        
        # Create mappings for overdue data
        file_to_overdue = {}
        file_to_critical = {}
        for file_result in overdue_data['file_results']:
            file_to_overdue[file_result['file_name']] = file_result['overdue_jobs_count']
            file_to_critical[file_result['file_name']] = file_result['critical_overdue_jobs_count']
        
        # Create data for overdue and critical overdue by date
        dates = []
        overdue_by_date = []
        critical_by_date = []
        
        for date, files in date_to_file.items():
            date_overdue = 0
            date_critical = 0
            
            for file in files:
                # Try exact match
                if file in file_to_overdue:
                    date_overdue += file_to_overdue[file]
                    date_critical += file_to_critical[file]
                else:
                    # Try basename
                    basename = os.path.basename(file)
                    if basename in file_to_overdue:
                        date_overdue += file_to_overdue[basename]
                        date_critical += file_to_critical[basename]
                    else:
                        # Try partial matching
                        for analysis_file in file_to_overdue:
                            if file in analysis_file or analysis_file in file:
                                date_overdue += file_to_overdue[analysis_file]
                                date_critical += file_to_critical[analysis_file]
                                break
            
            dates.append(date)
            overdue_by_date.append(date_overdue)
            critical_by_date.append(date_critical)
        
        # Sort the data by date
        sorted_indices = [i for i, _ in sorted(enumerate(dates), key=lambda x: x[1])]
        sorted_dates = [dates[i] for i in sorted_indices]
        sorted_overdue = [overdue_by_date[i] for i in sorted_indices]
        sorted_critical = [critical_by_date[i] for i in sorted_indices]
        
        # Add overdue jobs line
        if any(sorted_overdue):
            fig.add_trace(go.Scatter(
                x=sorted_dates,
                y=sorted_overdue,
                name='Overdue Jobs',
                line=dict(color='#FF9800', width=2, dash='dot'),  # Orange for Overdue Jobs
                mode='lines+markers+text',
                text=sorted_overdue,
                textposition="top center"
            ))
        
        # Add critical overdue jobs line
        if any(sorted_critical):
            fig.add_trace(go.Scatter(
                x=sorted_dates,
                y=sorted_critical,
                name='Critical Overdue',
                line=dict(color='#F44336', width=2, dash='dot'),  # Red for Critical Overdue
                mode='lines+markers+text',
                text=sorted_critical,
                textposition="top right"
            ))
    
    fig.update_layout(
        title='Job Trends Over Time',
        xaxis_title='Date',
        yaxis_title='Number of Jobs',
        height=400,
        showlegend=True
    )
    return fig

def create_jobs_pie_chart(df, overdue_data=None):
    """Create a pie chart showing the proportion of job statuses.
    
    Args:
        df: DataFrame with job data
        overdue_data: Optional dictionary with overdue jobs data
    """
    # Calculate base metrics
    total_jobs = df['Total Count of Jobs'].sum()
    new_jobs = df['New Job Count'].sum()
    
    # Calculate overdue and critical overdue values from file-level data
    overdue_jobs = 0
    critical_overdue = 0
    
    if overdue_data and 'file_results' in overdue_data and overdue_data['file_results']:
        # Create mappings for overdue data
        file_to_overdue = {}
        file_to_critical = {}
        for file_result in overdue_data['file_results']:
            file_to_overdue[file_result['file_name']] = file_result['overdue_jobs_count']
            file_to_critical[file_result['file_name']] = file_result['critical_overdue_jobs_count']
        
        # Sum up overdue jobs for files in the current filtered data
        for _, row in df.iterrows():
            file_name = row['File Name']
            
            # Try exact match
            if file_name in file_to_overdue:
                overdue_jobs += file_to_overdue[file_name]
                critical_overdue += file_to_critical[file_name]
            else:
                # Try basename matching
                basename = os.path.basename(file_name)
                if basename in file_to_overdue:
                    overdue_jobs += file_to_overdue[basename]
                    critical_overdue += file_to_critical[basename]
                else:
                    # Try partial matching
                    for analysis_file in file_to_overdue:
                        if file_name in analysis_file or analysis_file in file_name:
                            overdue_jobs += file_to_overdue[analysis_file]
                            critical_overdue += file_to_critical[analysis_file]
                            break
    
    # Calculate remaining jobs (total - new - overdue)
    # Note: overdue jobs might overlap with new jobs, so we need to be careful
    remaining_jobs = total_jobs - new_jobs - overdue_jobs
    if remaining_jobs < 0:
        remaining_jobs = 0
    
    # Prepare data for pie chart
    labels = []
    values = []
    colors = []
    
    if new_jobs > 0:
        labels.append('New Jobs')
        values.append(new_jobs)
        colors.append('#4CAF50')  # Green
    
    if overdue_jobs > 0:
        labels.append('Overdue Jobs')
        values.append(overdue_jobs)
        colors.append('#FF9800')  # Orange
    
    if critical_overdue > 0:
        labels.append('Critical Overdue')
        values.append(critical_overdue)
        colors.append('#F44336')  # Red
    
    if remaining_jobs > 0:
        labels.append('Other Jobs')
        values.append(remaining_jobs)
        colors.append('#1E88E5')  # Blue
    
    # Create pie chart
    if not labels:
        # Fallback if no data
        labels = ['No Data']
        values = [1]
        colors = ['#E0E0E0']
    
    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=.4,
        marker_colors=colors
    )])
    
    fig.update_layout(
        title='Job Status Distribution',
        height=400,
        showlegend=True
    )
    
    return fig

def create_overdue_jobs_chart(overdue_data, critical_data):
    """Create a bar chart comparing overdue and critical overdue jobs."""
    labels = ['Overdue Jobs', 'Critical Overdue Jobs']
    values = [overdue_data, critical_data]
    
    fig = go.Figure(data=[
        go.Bar(name='Count', x=labels, y=values, 
               marker_color=['#FF9800', '#F44336'])  # Orange for Overdue, Red for Critical
    ])
    
    fig.update_layout(
        title='Overdue Jobs Analysis',
        xaxis_title='Job Type',
        yaxis_title='Count',
        height=400,
        showlegend=False
    )
    
    return fig

def create_excel_report(df, analysis_results):
    """Create a formatted Excel report with job status data."""
    output = BytesIO()
    
    # Create a copy of the dataframe for the report
    report_df = df.copy()
    
    # Add overdue data if available
    if analysis_results and 'file_results' in analysis_results and analysis_results['file_results']:
        # Create mapping of file names to their analysis results
        file_analysis_map = {}
        for file_result in analysis_results['file_results']:
            file_name = file_result['file_name']
            file_analysis_map[file_name] = {
                'Overdue Jobs': file_result['overdue_jobs_count'],
                'Critical Overdue': file_result['critical_overdue_jobs_count'],
                'Overdue %': f"{file_result['overdue_jobs_percentage']}%",
                'Critical %': f"{file_result['critical_overdue_jobs_percentage']}%"
            }
        
        # Create simplified mapping for easier matching
        simplified_map = {}
        for full_path in file_analysis_map:
            simple_name = os.path.basename(full_path)
            simplified_map[simple_name] = file_analysis_map[full_path]
            simplified_map[full_path] = file_analysis_map[full_path]
        
        # Add overdue metrics to the dataframe
        overdue_jobs = []
        critical_overdue = []
        overdue_pct = []
        critical_pct = []
        
        for _, row in report_df.iterrows():
            file_name = row['File Name']
            simple_name = os.path.basename(file_name)
            
            if file_name in simplified_map:
                overdue_jobs.append(simplified_map[file_name]['Overdue Jobs'])
                critical_overdue.append(simplified_map[file_name]['Critical Overdue'])
                overdue_pct.append(simplified_map[file_name]['Overdue %'])
                critical_pct.append(simplified_map[file_name]['Critical %'])
            elif simple_name in simplified_map:
                overdue_jobs.append(simplified_map[simple_name]['Overdue Jobs'])
                critical_overdue.append(simplified_map[simple_name]['Critical Overdue'])
                overdue_pct.append(simplified_map[simple_name]['Overdue %'])
                critical_pct.append(simplified_map[simple_name]['Critical %'])
            else:
                # Try partial matching
                matched = False
                for analysis_file in file_analysis_map:
                    if file_name in analysis_file or analysis_file in file_name:
                        overdue_jobs.append(file_analysis_map[analysis_file]['Overdue Jobs'])
                        critical_overdue.append(file_analysis_map[analysis_file]['Critical Overdue'])
                        overdue_pct.append(file_analysis_map[analysis_file]['Overdue %'])
                        critical_pct.append(file_analysis_map[analysis_file]['Critical %'])
                        matched = True
                        break
                
                if not matched:
                    overdue_jobs.append("N/A")
                    critical_overdue.append("N/A")
                    overdue_pct.append("N/A")
                    critical_pct.append("N/A")
        
        # Add the overdue metrics to the dataframe
        report_df['Overdue Jobs'] = overdue_jobs
        report_df['Critical Overdue'] = critical_overdue
        report_df['Overdue %'] = overdue_pct
        report_df['Critical %'] = critical_pct
    else:
        # Add placeholders if no overdue analysis is available
        report_df['Overdue Jobs'] = "N/A"
        report_df['Critical Overdue'] = "N/A"
        report_df['Overdue %'] = "N/A"
        report_df['Critical %'] = "N/A"
    
    # Reorder columns for the report
    column_order = [
        'Date Extracted from File Name',
        'File Name', 
        'Vessel Name', 
        'Total Count of Jobs', 
        'New Job Count',
        'Overdue Jobs',
        'Critical Overdue',
        'Overdue %',
        'Critical %'
    ]
    
    # Only include columns that exist
    available_columns = [col for col in column_order if col in report_df.columns]
    report_df = report_df[available_columns]
    
    # Write to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report_df.to_excel(writer, sheet_name='Job Status Summary', index=False)
        
        # Get the workbook and worksheet to apply formatting
        workbook = writer.book
        worksheet = writer.sheets['Job Status Summary']
        
        # Apply basic formatting
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Header formatting
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output
